import configparser
import openpyxl as op
from datetime import datetime
from openpyxl.styles import Alignment
from random import randint
from lib_fetcher_image import FetcherImage

filename_output_xlsx = 'output.xlsx'
filename_cfg_ini = "task.ini"

class Bot():
	def __init__(self,namefile:str):
		self.cfg = configparser.ConfigParser()
		self.cfg.read(namefile, encoding='utf-8')
		self.fetcher = FetcherImage()

		self.headers = self.cfg['TaskInsertHeaders']['titles'].split("\n")
		self.text = self.cfg['TaskInsertText']['Text']
		self.company_info = self.cfg['TaskInsertCompanyInfo']
		self.date_time = self.cfg['TaskInsertDateTime']
		if self.date_time['value'] == "now":
			self.date_time['value'] = str(datetime.now().strftime('%d/%m/%Y %H:%M'))
		self.price_value = self.cfg['TaskInsertPrice']
		self.date_time_full = f"{self.date_time['value']}"
		self.precet = self.cfg['TaskInsertPrecet']

	def iterator_rows(self):
		for header in self.headers:
			temp_rows = []
			temp_rows.append(self.company_info['email'])
			temp_rows.append(self.date_time_full)
			temp_rows.append(header)
			temp_rows.append(self.text.replace("ARTICLE_CODE", self.generate_article_code()))
			temp_rows.append(self.price_value['value'])
			temp_rows.append(" ".join(self.fetcher.get_images_urls(header, 2)))
			temp_rows.append(self.precet['name'])
			temp_rows.append(self.precet['city'])
			temp_rows.append(self.company_info['phone'])
			temp_rows.append(self.company_info['id_package'])
			yield temp_rows

	def generate_article_code(self):
		return f"ARTICLE{randint(111111,999999)}"

	def write_row(self, sheet, counter, row):
		a,b,c,d,e,f,g,h,i,j = row
		sheet[f"A{counter}"] = a 
		sheet[f"B{counter}"] = b
		sheet[f"C{counter}"] = c
		sheet[f"D{counter}"] = d 
		sheet[f"E{counter}"] = e
		sheet[f"F{counter}"] = f
		sheet[f"G{counter}"] = g 
		sheet[f"H{counter}"] = h
		sheet[f"I{counter}"] = i
		sheet[f"J{counter}"] = j
		sheet[f"A{counter}"].alignment = Alignment(wrapText=True) 
		sheet[f"B{counter}"].alignment = Alignment(wrapText=True)
		sheet[f"C{counter}"].alignment = Alignment(wrapText=True)
		sheet[f"D{counter}"].alignment = Alignment(wrapText=True) 
		sheet[f"E{counter}"].alignment = Alignment(wrapText=True)
		sheet[f"F{counter}"].alignment = Alignment(wrapText=True)
		sheet[f"G{counter}"].alignment = Alignment(wrapText=True) 
		sheet[f"H{counter}"].alignment = Alignment(wrapText=True)
		sheet[f"I{counter}"].alignment = Alignment(wrapText=True)
		sheet[f"J{counter}"].alignment = Alignment(wrapText=True)


	def write_xlsx(self):
		list_tab = [
				"Аккаунт* (логин)", 
				'Дата и время публикации *(формат ДД.ММ.ГГГГ ЧЧ:ММ',
				"Заголовок*",
				"Текст*",
				"Цена*",
				"Пресет фото*, либо путь к ресурсу, инструкция: https://app.ayicrm.ru/#/help/37",
				"Пресет параметров* (название пресета)",
				"Пресет мест *(название пресета)",
				"Телефон объявления",
				"ID пакета (для авито)",
			]

		excel_doc = op.Workbook()
		excel_doc.create_sheet(title = 'Лист1', index = 0)
		sheetnames = excel_doc.sheetnames 
		sheet = excel_doc[sheetnames[0]]
		counter = 1
		self.write_row(sheet, counter, list_tab)
		counter += 1
		for row in self.iterator_rows():
			self.write_row(sheet, counter, row)
			counter += 1

		excel_doc.save(filename_output_xlsx)

pt = Bot(filename_cfg_ini)
pt.write_xlsx()